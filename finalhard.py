from statistics import mode
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()
ab = Workbook()
cb = Workbook()


wb = load_workbook("D:\\xlfile\listitems.xlsx")
tot_price = 0
a = wb.active

j=1
print("WELCOME")
while j<1000:
    
    name=input("ENTER THE PRODUCT : ")
    if name=="N":
        break
    else:
        quantities=int(input("ENTER THE QUANTITIES : "))
        for i in range(1,100000):
            product = a.cell(row=i,column=2)
            if product.value == name:
                quantity = a.cell(row=i,column=3).value
                
                a.cell(row=i,column=3,value=(quantity-quantities))
                
                print("ADD ITEMS TO YOUR CART \n OTHERWISE")
                print("PRESS N TO EXIT")
                tot_price += (a.cell(row=i,column=4).value)*quantities
                wb.save("D:\\xlfile\listitems.xlsx")
                ab=load_workbook("soldout.xlsx")
                b = ab.active
                for k in range(1,100000):
                    product = b.cell(row=k,column=2)
                    if product.value == name:
                        quantity = b.cell(row=k,column=3).value
                        b.cell(row=i,column=3,value=(quantity+quantities))
                        ab.save("soldout.xlsx")
                        break
                    else:
                        pass
                break
print('TOTAL PRICE = ',tot_price)
for i in range(2,141):
    if int ((a.cell(row=i,column=3).value)<=10):
    
       
        aircel = a.cell(row=i,column=2).value
        email=a.cell(row=i,column=5).value
        
        from email.message import EmailMessage
        import ssl
        import smtplib

        email_sender = "jrn03540@gmail.com"
        email_password = "jimzlkimpykmlaqh"
        email_receiver = email
        subject = "This product is below the limit.So, please send the stock ASAP!"
        body =aircel
        en = EmailMessage()
        en['From'] = email_sender
        en['To'] = email_receiver
        en['subject'] = subject
        en.set_content(body)

        context = ssl.create_default_context()

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
            smtp.login(email_sender,email_password)
            smtp.sendmail(email_sender,email_receiver,en.as_string())
            
        
        
    else:
        pass

cb=load_workbook("soldout.xlsx")
c=cb.active
max=c.cell(row=2,column=3).value
demand=c.cell(row=2,column=2).value

for l in range(2,141):
    if int ((c.cell(row=l,column=3).value)>max):
        demand=c.cell(row=l,column=2).value
print("THIS PRODUCT DEMAND IS HIGH:",demand)
cb.save("soldout.xlsx")