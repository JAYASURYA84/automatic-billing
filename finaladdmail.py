from statistics import mode
from openpyxl import Workbook
from openpyxl import load_workbook
wb = Workbook()


wb = load_workbook("D:\\xlfile\listitems.xlsx")
tot_price = 0
a = wb.active
j=1
print("WELCOME")
while j<1000:
    
    name=input("ENTER THE PRODUCT : ")
    if name=="N":
        break
    else:
        quantities=int(input("ENTER THE QUANTITIES : "))
        for i in range(1,100000):
            product = a.cell(row=i,column=2)
            if product.value == name:
                quantity = a.cell(row=i,column=3).value
                
                a.cell(row=i,column=3,value=(quantity-quantities))
                print("ADD ITEMS TO YOUR CART \n OTHERWISE")
                print("PRESS N TO EXIT")
                tot_price += (a.cell(row=i,column=4).value)*quantities
                break
print('TOTAL PRICE = ',tot_price)
for i in range(2,16):
    if int (a.cell(row=i,column=3).value)<=10:
    
        print("This product is below the limit:",a.cell(row=i,column=2).value)
        aircel = a.cell(row=i,column=2).value
        wb.save("D:\\xlfile\listitems.xlsx")
        from email.message import EmailMessage
        import ssl
        import smtplib

        email_sender = "jrn03540@gmail.com"
        email_password = "jimzlkimpykmlaqh"
        if (aircel=="Salt"):
            email_receiver = "lioneljames987@gmail.com"
            subject = "surya"
            body =aircel
            en = EmailMessage()
            en['From'] = email_sender
            en['To'] = email_receiver
            en['subject'] = subject
            en.set_content(body)

            context = ssl.create_default_context()

            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
                smtp.login(email_sender,email_password)
                smtp.sendmail(email_sender,email_receiver,en.as_string())
        elif(aircel=="Garlic"):
             email_receiver = "suryasurya12837@gmail.com"
             subject = "surya"
             body =aircel
             en = EmailMessage()
             en['From'] = email_sender
             en['To'] = email_receiver
             en['subject'] = subject
             en.set_content(body)

             context = ssl.create_default_context()

             with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
                smtp.login(email_sender,email_password)
                smtp.sendmail(email_sender,email_receiver,en.as_string())
        elif(aircel=="Pepper"):
             email_receiver = "thulasidharanb26@gmail.com"
             subject = "surya"
             body =aircel
             en = EmailMessage()
             en['From'] = email_sender
             en['To'] = email_receiver
             en['subject'] = subject
             en.set_content(body)

             context = ssl.create_default_context()

             with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
                smtp.login(email_sender,email_password)
                smtp.sendmail(email_sender,email_receiver,en.as_string())
        else:
            pass
    else:
        pass
wb.save("D:\\xlfile\listitems.xlsx")